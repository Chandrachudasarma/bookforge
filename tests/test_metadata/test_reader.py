"""Tests for the Excel metadata reader.

Creates real .xlsx files via openpyxl for testing.
"""

from pathlib import Path

import pytest
import openpyxl

from bookforge.core.exceptions import MetadataError
from bookforge.metadata.reader import read_metadata


# ---------------------------------------------------------------------------
# Fixtures — create real Excel files
# ---------------------------------------------------------------------------


@pytest.fixture
def simple_excel(tmp_path: Path) -> Path:
    """Excel with standard column headers and 2 data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title", "Author", "ISBN", "Year", "Language", "Output Formats"])
    ws.append(["Book One", "Dr. John Smith", "978-0-1234-5678-0", 2026, "en", "epub,docx"])
    ws.append(["Book Two", "Prof. Jane Doe", "978-0-9876-5432-1", 2025, "en", "pdf"])

    path = tmp_path / "metadata.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_with_blank_rows(tmp_path: Path) -> Path:
    """Excel with blank rows interspersed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title", "Author"])
    ws.append(["Book One", "Author One"])
    ws.append([None, None])  # blank row
    ws.append(["Book Two", "Author Two"])

    path = tmp_path / "blanks.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def excel_with_chapter_order(tmp_path: Path) -> Path:
    """Excel with Chapter Order and Input Files columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title", "Author", "Input Files", "Chapter Order"])
    ws.append(["My Book", "Author", "intro.html", 1])
    ws.append(["My Book", "Author", "methods.html", 2])
    ws.append(["My Book", "Author", "results.html", 3])

    path = tmp_path / "ordered.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def columns_config() -> dict:
    """Standard column mapping matching columns.yaml."""
    return {
        "title": "Title",
        "author_name": "Author",
        "isbn": "ISBN",
        "year": "Year",
        "language": "Language",
        "output_formats": "Output Formats",
        "input_files": "Input Files",
        "chapter_order": "Chapter Order",
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_reads_simple_excel(simple_excel, columns_config):
    rows = read_metadata(simple_excel, columns_config)
    assert len(rows) == 2
    assert rows[0]["title"] == "Book One"
    assert rows[1]["title"] == "Book Two"


def test_maps_columns_to_canonical_names(simple_excel, columns_config):
    rows = read_metadata(simple_excel, columns_config)
    assert rows[0]["author_name"] == "Dr. John Smith"
    assert rows[0]["isbn"] == "978-0-1234-5678-0"
    assert rows[0]["year"] == 2026


def test_skips_blank_rows(excel_with_blank_rows, columns_config):
    rows = read_metadata(excel_with_blank_rows, columns_config)
    assert len(rows) == 2
    assert rows[0]["title"] == "Book One"
    assert rows[1]["title"] == "Book Two"


def test_includes_row_index(simple_excel, columns_config):
    rows = read_metadata(simple_excel, columns_config)
    assert "_row_index" in rows[0]
    assert rows[0]["_row_index"] == 1
    assert rows[1]["_row_index"] == 2


def test_reads_chapter_order(excel_with_chapter_order, columns_config):
    rows = read_metadata(excel_with_chapter_order, columns_config)
    assert len(rows) == 3
    assert rows[0]["input_files"] == "intro.html"
    assert rows[0]["chapter_order"] == 1
    assert rows[2]["chapter_order"] == 3


def test_raises_on_nonexistent_file(columns_config):
    with pytest.raises(MetadataError, match="Cannot open Excel"):
        read_metadata(Path("/nonexistent/file.xlsx"), columns_config)


def test_raises_on_empty_data(tmp_path, columns_config):
    """Excel with headers only, no data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title", "Author"])
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))

    with pytest.raises(MetadataError, match="no data rows"):
        read_metadata(path, columns_config)


def test_handles_unmapped_columns(tmp_path, columns_config):
    """Excel with columns not in the mapping — they're ignored."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title", "Author", "Random Extra Column"])
    ws.append(["Book", "Writer", "ignored value"])
    path = tmp_path / "extra.xlsx"
    wb.save(str(path))

    rows = read_metadata(path, columns_config)
    assert len(rows) == 1
    assert rows[0]["title"] == "Book"
    assert "Random Extra Column" not in rows[0]
