"""Excel metadata reader — reads book metadata from .xlsx files.

Applies column mapping from config/columns.yaml so that the Excel sheet
can use any column headers — the mapping translates them to canonical
field names used by the pipeline.

Each row in the Excel sheet represents one book (or one book in a batch).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import yaml

from bookforge.core.exceptions import MetadataError
from bookforge.core.logging import get_logger

logger = get_logger("metadata.reader")


def read_metadata(
    excel_path: Path,
    columns_config: dict | None = None,
) -> list[dict]:
    """Read Excel file and return a list of raw metadata row dicts.

    Args:
        excel_path:     Path to the .xlsx file.
        columns_config: Mapping of canonical_name → Excel column header.
                        If None, loads from config/columns.yaml.

    Returns:
        List of dicts, one per non-blank row, keyed by canonical field names.

    Raises:
        MetadataError: If the Excel file cannot be read or has no data rows.
    """
    if columns_config is None:
        columns_config = load_columns_config()

    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    except Exception as exc:
        raise MetadataError(f"Cannot open Excel file: {excel_path.name}: {exc}") from exc

    ws = wb.active
    if ws is None:
        wb.close()
        raise MetadataError(f"Excel file has no active worksheet: {excel_path.name}")

    # Read headers from row 1
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        wb.close()
        raise MetadataError(f"Excel file has no header row: {excel_path.name}")

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

    # Build reverse mapping: Excel column name → canonical name
    col_map: dict[str, str] = {v: k for k, v in columns_config.items()}

    # Map column indices to canonical names
    col_indices: dict[int, str] = {}
    for i, header in enumerate(headers):
        if header in col_map:
            col_indices[i] = col_map[header]

    if not col_indices:
        wb.close()
        logger.warning(
            "No Excel columns matched the column mapping",
            excel_headers=headers,
            expected=list(columns_config.values()),
        )

    # Read data rows
    rows: list[dict] = []
    row_num = 1  # 0-based index for data rows (row 2 in Excel = index 0)

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict: dict = {}
        for i, value in enumerate(row):
            if i in col_indices:
                row_dict[col_indices[i]] = value

        # Skip completely blank rows
        if any(v is not None for v in row_dict.values()):
            row_dict["_row_index"] = row_num
            rows.append(row_dict)

        row_num += 1

    wb.close()

    if not rows:
        raise MetadataError(f"Excel file has no data rows: {excel_path.name}")

    logger.debug("Excel metadata read", rows=len(rows), columns=list(col_indices.values()))
    return rows


def load_columns_config(config_path: Path | None = None) -> dict:
    """Load column mapping from config/columns.yaml.

    Resolves the path relative to the project root (directory containing
    pyproject.toml), not CWD.

    Returns:
        Dict of canonical_name → Excel column header string.
    """
    if config_path is None:
        config_path = _find_project_file("config/columns.yaml")

    if not config_path.exists():
        raise MetadataError(f"Columns config not found: {config_path}")

    try:
        data = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise MetadataError(f"Invalid columns config: {exc}") from exc

    mappings = data.get("mappings")
    if not mappings or not isinstance(mappings, dict):
        raise MetadataError("columns.yaml must contain a 'mappings' dict")

    return mappings


def _find_project_file(relative_path: str) -> Path:
    """Resolve a path relative to the project root (dir containing pyproject.toml)."""
    cwd = Path.cwd()
    for parent in [cwd, *cwd.parents]:
        if (parent / "pyproject.toml").exists():
            return parent / relative_path
    return cwd / relative_path
